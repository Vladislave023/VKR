from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook


HEADER_ROWS = {
    "школы",
    "учебные подразделения",
    "двфу",
    "анализ штатного расписания",
}


def _clean(value: Any) -> str:
    """Convert a worksheet value to normalized text."""
    if value is None:
        return ""
    return " ".join(str(value).split())


def _is_school_row(value: str) -> bool:
    """Return True when the first column value looks like a school/institute row."""
    lower = value.lower()
    if lower in HEADER_ROWS:
        return False
    if "школа" not in lower and "институт" not in lower:
        return False
    return not any(char.isdigit() for char in value[:8])


def _is_department_row(value: str) -> bool:
    """Return True when the first column value looks like a department row."""
    lower = value.lower()
    return "кафедр" in lower or "департамент" in lower


def parse_sr(filepath: str) -> tuple[dict[str, list[str]], list[str]]:
    """
    Parse the staff allocation workbook.

    Returns ``(school_department_map, errors)`` where the map contains
    school names as keys and ordered, de-duplicated department names as values.
    """
    result: dict[str, list[str]] = {}
    errors: list[str] = []
    current_school: str | None = None

    try:
        workbook = load_workbook(Path(filepath), read_only=True, data_only=True)
    except Exception as exc:
        return {}, [f"Не удалось открыть файл штатной расстановки: {exc}"]

    worksheet = workbook.worksheets[0]

    for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        value = _clean(row[0] if row else None)
        if not value:
            continue

        if _is_school_row(value):
            current_school = value
            result.setdefault(current_school, [])
            continue

        if not _is_department_row(value):
            continue

        if not current_school:
            errors.append(f"Строка {row_number}: кафедра/департамент без школы: {value}")
            continue

        departments = result.setdefault(current_school, [])
        if value not in departments:
            departments.append(value)

    workbook.close()
    return result, errors
