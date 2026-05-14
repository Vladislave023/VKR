from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


SPECIALTY_CODE_RE = re.compile(r"^\d{2}\.\d{2}\.\d{2}$")


def _clean(value: Any) -> str:
    """Convert a worksheet value to normalized text."""
    if value is None:
        return ""
    return " ".join(str(value).split())


def _normalize_level(value: Any) -> str | None:
    """Return canonical education level from a КУГ level marker."""
    text = _clean(value).lower()
    if "бакалавриат" in text:
        return "Бакалавриат"
    if "магистратура" in text:
        return "Магистратура"
    if "специалитет" in text:
        return "Специалитет"
    if "аспирантура" in text:
        return "Аспирантура"
    return None


def _is_course(value: Any) -> bool:
    """Check whether a value looks like an integer course number."""
    if isinstance(value, bool):
        return False
    if isinstance(value, int):
        return True
    if isinstance(value, float):
        return value.is_integer()
    return False


def parse_kug(filepath: str) -> tuple[list[dict[str, str]], list[str]]:
    """
    Parse the multi-sheet summary КУГ workbook.

    Returns a tuple ``(records, errors)``. Each record is unique by
    ``(institute_name, education_level, specialty_code)`` and contains:
    ``institute_name``, ``sheet_abbr``, ``education_level``,
    ``specialty_code`` and ``specialty_name``.
    """
    records: list[dict[str, str]] = []
    errors: list[str] = []
    seen: set[tuple[str, str, str]] = set()

    try:
        workbook = load_workbook(Path(filepath), read_only=True, data_only=True)
    except Exception as exc:
        return [], [f"Не удалось открыть файл КУГ: {exc}"]

    for worksheet in workbook.worksheets:
        if worksheet.title.strip().lower().startswith("титул"):
            continue

        institute_name = _clean(worksheet.cell(row=1, column=1).value)
        if not institute_name:
            errors.append(f"Лист {worksheet.title}: не найдено название школы в A1")
            continue

        current_level: str | None = None
        unsupported_level = False

        for row_number, row in enumerate(worksheet.iter_rows(min_row=11, values_only=True), start=11):
            values = list(row[:6])
            first = values[0] if len(values) > 0 else None
            level = _normalize_level(first)
            if level:
                current_level = level
                unsupported_level = False
                continue
            if "образование" in _clean(first).lower() and not level:
                current_level = None
                unsupported_level = True
                continue

            if _clean(first) == "< >":
                continue

            if not _is_course(first):
                continue

            try:
                specialty_code = _clean(values[2])
                specialty_name = _clean(values[3])

                if not current_level:
                    if unsupported_level:
                        continue
                    errors.append(
                        f"Лист {worksheet.title}, строка {row_number}: строка данных без маркера уровня"
                    )
                    continue

                if not SPECIALTY_CODE_RE.match(specialty_code):
                    errors.append(
                        f"Лист {worksheet.title}, строка {row_number}: некорректный код направления {specialty_code!r}"
                    )
                    continue

                if not specialty_name:
                    errors.append(
                        f"Лист {worksheet.title}, строка {row_number}: пустое название направления"
                    )
                    continue

                key = (institute_name, current_level, specialty_code)
                if key in seen:
                    continue

                seen.add(key)
                records.append(
                    {
                        "institute_name": institute_name,
                        "sheet_abbr": worksheet.title,
                        "education_level": current_level,
                        "specialty_code": specialty_code,
                        "specialty_name": specialty_name,
                    }
                )
            except Exception as exc:
                errors.append(f"Лист {worksheet.title}, строка {row_number}: ошибка парсинга: {exc}")

    workbook.close()
    return records, errors
